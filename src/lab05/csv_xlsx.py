from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv,sys
from pathlib import Path
sys.path.append('C:/Users/dasha/Desktop/python_labs/src')
from lab04.io_txt_csv import ensure_parent_dir

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    '''Функция конвертирует CSV-файл в XSLX-файл, проверяя синтаксис и корректность входного и создавая директорию(если надо) выходного'''
    c_path=Path(csv_path)
    if not c_path.exists():
        raise FileNotFoundError('Файл не найден')
    if c_path.suffix != '.csv':
            raise ValueError("Неверный тип файла")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    with open(c_path,'r',encoding='utf-8') as c_file:
        c_data=csv.DictReader(c_file)
        if not c_data.fieldnames:
            raise ValueError('Файл пустой или в нем нет заголовков')     
        ws.append(c_data.fieldnames)
        for row in c_data:
            ws.append([row[field] for field in c_data.fieldnames])

    x_path=Path(xlsx_path)
    ensure_parent_dir(x_path)
    for column in ws.columns:
            max_length=8
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                max_length = max(len(str(cell.value)), max_length)
            ws.column_dimensions[column_letter].width = max_length
    wb.save(x_path)
